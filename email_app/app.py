#!/usr/bin/env python3
"""
Professional AWS SES mailer for Joblynk.

Features:
- Single send and bulk send
- CSV/XLSX recipient ingestion
- Personalization tokens
- Batch + interval throttling
- Retry with exponential backoff
- Structured logging
- Environment-based SMTP credentials
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import smtplib
import ssl
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Iterable, List, Optional

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


@dataclass
class Recipient:
    full_name: str
    email: str
    phone: str = ""

    @property
    def first_name(self) -> str:
        return (self.full_name or "").strip().split(" ")[0] if self.full_name else "there"


class SesMailer:
    def __init__(self, host: str, port: int, username: str, password: str, timeout: int = 30):
        self.host = host
        self.port = port
        self.username = username
        self.password = password
        self.timeout = timeout

    def send_html(self, from_email: str, to_email: str, subject: str, html: str, reply_to: Optional[str] = None) -> None:
        msg = MIMEMultipart("alternative")
        msg["From"] = from_email
        msg["To"] = to_email
        msg["Subject"] = subject
        if reply_to:
            msg["Reply-To"] = reply_to

        msg.attach(MIMEText(html, "html", "utf-8"))

        context = ssl.create_default_context()
        with smtplib.SMTP(self.host, self.port, timeout=self.timeout) as server:
            server.ehlo()
            server.starttls(context=context)
            server.ehlo()
            server.login(self.username, self.password)
            server.sendmail(from_email, [to_email], msg.as_string())


def load_template(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"Template not found: {path}")
    return path.read_text(encoding="utf-8")


def read_recipients(file_path: Path) -> List[Recipient]:
    ext = file_path.suffix.lower()
    if ext == ".csv":
        return _read_csv(file_path)
    if ext == ".xlsx":
        return _read_xlsx(file_path)
    raise ValueError("Recipients file must be .csv or .xlsx")


def _read_csv(file_path: Path) -> List[Recipient]:
    rows: List[Recipient] = []
    with file_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            full_name = (row.get("Full Name") or row.get("full_name") or "").strip()
            email = (row.get("Email") or row.get("email") or "").strip()
            phone = (row.get("Phone Number") or row.get("phone") or "").strip()
            if full_name and email and EMAIL_RE.match(email):
                rows.append(Recipient(full_name=full_name, email=email, phone=phone))
    return rows


def _read_xlsx(file_path: Path) -> List[Recipient]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for .xlsx support. Install with: pip install openpyxl")

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}

    def get(row, key):
        i = idx.get(key)
        if i is None or i >= len(row):
            return ""
        v = row[i].value
        return str(v).strip() if v is not None else ""

    rows: List[Recipient] = []
    for r in ws.iter_rows(min_row=2):
        full_name = get(r, "Full Name")
        email = get(r, "Email")
        phone = get(r, "Phone Number")
        if full_name and email and EMAIL_RE.match(email):
            rows.append(Recipient(full_name=full_name, email=email, phone=phone))
    return rows


def personalize(template: str, recipient: Recipient) -> str:
    return (
        template
        .replace("{{full_name}}", recipient.full_name)
        .replace("{{first_name}}", recipient.first_name)
        .replace("{{email}}", recipient.email)
        .replace("{{phone}}", recipient.phone)
    )


def log_line(path: Path, message: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    with path.open("a", encoding="utf-8") as f:
        f.write(f"[{ts}] {message}\n")


def send_with_retry(mailer: SesMailer, from_email: str, recipient: Recipient, subject: str, html: str, reply_to: Optional[str], retries: int, base_backoff: float) -> None:
    attempt = 0
    while True:
        try:
            mailer.send_html(from_email, recipient.email, subject, html, reply_to=reply_to)
            return
        except Exception:
            attempt += 1
            if attempt > retries:
                raise
            time.sleep(base_backoff * (2 ** (attempt - 1)))


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Joblynk AWS SES mailer")
    p.add_argument("--mode", choices=["single", "bulk"], default="single")
    p.add_argument("--to", default="ryzsale@gmail.com", help="single mode recipient email")
    p.add_argument("--name", default="Riaz Mohammed", help="single mode recipient name")
    p.add_argument("--phone", default="+17732739855", help="single mode recipient phone")
    p.add_argument("--recipients", default="recipients.xlsx", help="bulk recipients file (.xlsx or .csv)")
    p.add_argument("--template", default="template.html", help="HTML template path")
    p.add_argument("--subject", default="Introduction to Joblynk Team")
    p.add_argument("--batch-size", type=int, default=100)
    p.add_argument("--batch-interval", type=float, default=60.0)
    p.add_argument("--per-email-delay", type=float, default=0.25)
    p.add_argument("--retries", type=int, default=3)
    p.add_argument("--timeout", type=int, default=30)
    p.add_argument("--dry-run", action="store_true")
    return p.parse_args()


def require_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def main() -> int:
    args = parse_args()

    smtp_host = os.getenv("SES_SMTP_HOST", "email-smtp.us-east-2.amazonaws.com")
    smtp_port = int(os.getenv("SES_SMTP_PORT", "587"))
    smtp_user = require_env("SES_SMTP_USERNAME")
    smtp_pass = require_env("SES_SMTP_PASSWORD")
    from_email = os.getenv("SES_FROM_EMAIL", "info@joblynk.ai")
    reply_to = os.getenv("SES_REPLY_TO", "adam@joblynk.ai")

    template = load_template(Path(args.template))
    log_path = Path("logs") / f"send-{datetime.now(timezone.utc).strftime('%Y%m%d')}.log"

    if args.mode == "single":
        recipients = [Recipient(full_name=args.name, email=args.to, phone=args.phone)]
    else:
        recipients = read_recipients(Path(args.recipients))

    if not recipients:
        print("No valid recipients found.")
        return 1

    print(f"Mode: {args.mode} | Recipients: {len(recipients)} | Dry run: {args.dry_run}")

    mailer = SesMailer(smtp_host, smtp_port, smtp_user, smtp_pass, timeout=args.timeout)

    sent = 0
    failed = 0

    for idx, recipient in enumerate(recipients, start=1):
        html = personalize(template, recipient)
        try:
            if args.dry_run:
                print(f"[DRY RUN] Would send to {recipient.email} ({recipient.full_name})")
                log_line(log_path, f"DRY_RUN OK {recipient.email} {recipient.full_name}")
            else:
                send_with_retry(
                    mailer=mailer,
                    from_email=from_email,
                    recipient=recipient,
                    subject=args.subject,
                    html=html,
                    reply_to=reply_to,
                    retries=args.retries,
                    base_backoff=1.5,
                )
                print(f"Sent: {recipient.email}")
                log_line(log_path, f"SENT OK {recipient.email} {recipient.full_name}")
            sent += 1
        except Exception as e:
            failed += 1
            print(f"Failed: {recipient.email} | {e}")
            log_line(log_path, f"SENT FAIL {recipient.email} {recipient.full_name} | {e}")

        if idx < len(recipients):
            time.sleep(args.per_email_delay)
            if args.mode == "bulk" and idx % args.batch_size == 0:
                print(f"Batch pause: {args.batch_interval}s")
                time.sleep(args.batch_interval)

    print(f"Done. Sent={sent} Failed={failed}")
    return 0 if failed == 0 else 2


if __name__ == "__main__":
    sys.exit(main())
